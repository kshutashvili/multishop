from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from django.utils.translation import get_language
from django.utils.translation import ugettext as _

from shop.catalogue.models import Product
from shop.catalogue.models import AttributeOption
from shop.catalogue.models import ProductAttributeValue
from .base import Base


class CatelogueExporter(Base):

    attributes_to_export = {}
    data_rows_count = None

    def __init__(self, data, *args, **kwargs):
        self.form_data = data

    def get_products_for_export(self, p_class):
        products = Product.objects.filter(product_class=p_class)
        return products

    def get_fields_to_export(self):
        for field in self.form_data['fields']:
            if field == 'partner_sku':
                self.FIELDS = self.FIELDS + ((field, unicode(_('SKU'))),)
            else:
                name = Product._meta.get_field(field).verbose_name
                self.FIELDS = self.FIELDS + ((field, unicode(name)),)
        return self.FIELDS

    def get_attributes_to_export(self):
        self.attributes_to_export = self.form_data['attributes']
        return self.attributes_to_export

    def handle(self):
        product_class = self.form_data['product_class']
        products = self.get_products_for_export(product_class)
        self.data_rows_count = len(products) + 3
        self.get_attributes_to_export()
        self.get_fields_to_export()
        result = self.export(products)
        return result

    def export(self, products):
        """
            Creates a xlsx file with products data
        """
        wb = Workbook()
        ws = wb.active
        ws = self.create_heading(ws)

        for i, product in enumerate(products):
            data = self.get_product_data(product)
            for j, value in enumerate(data):
                ws.cell(row=i + 4, column=j + 1, value=value)

        return wb

    def create_heading(self, ws):
        i = 1
        heading_fill = PatternFill("solid", fgColor="00ffc269")
        codes_fill = PatternFill("solid", fgColor="00ded6d6")
        danger_fill = PatternFill("solid", fgColor="00ff6969")

        for key, value in self.FIELDS:
            ws.cell(row=1, column=i, value=unicode(value)) \
                .fill = heading_fill
            ws.cell(row=2, column=i, value=key) \
                .fill = codes_fill
            i += 1

        colomn = i
        for attr in self.attributes_to_export:
            ws.cell(row=1, column=colomn, value=attr.name) \
                .fill = heading_fill
            ws.cell(row=2, column=colomn, value=attr.code) \
                .fill = codes_fill
            colomn += 1

        cell = ws.cell(row=1, column=colomn + 1, value=get_language().upper())
        cell.alignment = Alignment(horizontal='center')
        cell.fill = danger_fill

        self.adjust_colomn_width(ws)

        ws.row_dimensions[1].height = 20
        return ws

    def get_product_data(self, product):
        result = []

        for field_code, name in self.FIELDS:
            if field_code == 'categories':
                value = self.get_categories(product)
            elif field_code == 'partner_sku':
                value = self.get_sku(product)
            else:
                value = unicode(getattr(product, field_code))
            result.append(value)

        for attr in self.attributes_to_export:
            result.append(self.get_attribute_value(product, attr))

        return result

    def get_attribute_value(self, product, attribute):
        code = attribute.code
        try:
            value = product.attribute_values.get(attribute__code=code)
            if isinstance(value.value, tuple):
                if get_language() == 'ru':
                    result = value.value[0]
                else:
                    result = value.value[1]
            elif isinstance(value.value, AttributeOption):
                result = value.value.option
            else:
                result = value.value

        except ProductAttributeValue.DoesNotExist:
            result = None
        return result

    def get_categories(self, product):
        return self.format_comma(
            list(product.categories.all().values_list('id', flat=True))
        )

    def get_sku(self, product):
        return product.stockrecords.first().partner_sku

    def format_comma(self, categories):
        categories = [str(category_id) for category_id in categories]
        return ', '.join(categories)

    def adjust_colomn_width(self, worksheet):
        for column_cells in worksheet.columns:
            length = max(len(unicode(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column].width = length
