from io import BytesIO
from openpyxl import load_workbook
from django.utils import translation

from .base import Base
from shop.catalogue.models import Product
from shop.catalogue.models import Category
from shop.catalogue.models import ProductCategory
from shop.catalogue.models import AttributeOption
from shop.catalogue.models import ProductClass


class CatalogueImporter(Base):

    def __init__(self, file):
        self.wb = load_workbook(BytesIO(file.read()))
        self.ws = self.wb.active

    def handle(self):
        self.statistics = {
            'created': 0,
            'updated': 0,
            'errors': [],
        }
        self._import()
        return self.statistics

    def _import(self):
        self.max_row = self.ws.max_row
        self.get_codes()
        self.get_language()
        for row in self.ws:
            if row[0].row > 3:
                try:
                    self.create_update_product(row[:self.data_colomns_len])
                except():
                    self.statistics['errors'].append(str(row[0].row))

    def get_codes(self):
        rows = self.ws.rows
        self.codes = [i.value for i in list(rows)[1] if i.value is not None]
        self.data_colomns_len = len(self.codes)

    def get_language(self):
        position = self.data_colomns_len + 1
        self.lang = list(self.ws.rows)[0][position].value.lower()
        translation.activate(self.lang)

    def create_update_product(self, data):
        values = [item.value for item in data]
        values = dict(zip(self.codes, values))
        try:
            product = Product.objects.get(id=values['id'])
            self.statistics['updated'] += 1
        except Product.DoesNotExist:
            product = Product()
            p_class = ProductClass.objects.get(name=values['product_class'])
            product.product_class = p_class
            self.statistics['created'] += 1

        for code, value in values.iteritems():
            if code == 'product_class':
                _class = ProductClass.objects.get(name=values['product_class'])
                product.product_class = _class
            elif code == 'categories':
                categories = Category.objects.filter(
                    id__in=self._get_categories(values['categories']))
                ProductCategory.objects.filter(product=product).delete()
                for category in categories:
                    product_category = ProductCategory()
                    product_category.product = product
                    product_category.category = category
                    product_category._no_index = True
                    product_category.save()
            elif hasattr(product, code):
                setattr(product, code, value)
            elif product.attributes.filter(code=code).exists():
                p_value = product.attribute_values.get(attribute__code=code)
                if isinstance(p_value.value, tuple):
                    attribute = 'value_{}_{}'.format(
                        p_value.attribute.type,
                        self.lang,
                    )
                    setattr(p_value, attribute, value)
                else:
                    try:
                        p_value._set_value(value)
                    except AttributeOption.DoesNotExist:
                        attr_option = AttributeOption.objects.create(
                            group=p_value.attribute.option_group,
                            option=unicode(value)
                        )
                        p_value._set_value(attr_option)
                p_value.save()
        if not data[0].row == self.max_row:
            product._no_index = True
        product.save()
        return product

    def _get_categories(self, category):
        if isinstance(category, type(None)):
            categories = []
        else:
            category = unicode(category)
            categories = category.split(', ')
        return categories
